import os
import secrets
import sys
import psycopg2
import openpyxl

# --- CONFIGURATION ---
EXCEL_PATH = r"اسماء شركات الاسنان\جليانة_علاقات_دقيقة.xlsx"
JFZ_CODE = "JFZ"
DENTAL_CEILING = 3000.00
DENTAL_COV = 100.00
DB_URL = "postgresql://postgres:RJzgbxCvJmUgaFQmp6eTWzEwRbAzYDrz@localhost:5432/wahda_db"

def generate_cuid():
    # CUID-like string: 'c' + 24 lowercase hex characters
    return "c" + secrets.token_hex(12)

def main():
    apply = "--apply" in sys.argv

    print("=" * 60)
    print("📋  IMPORT GELYANA BENEFICIARIES FOR DENTAL (JFZ)")
    print(f"    Mode: {'🔴 APPLY CHANGES TO DATABASE' if apply else '🔵 DRY RUN (Simulation)'}")
    print("=" * 60)

    # 1. Verify Excel File
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel file not found at: {EXCEL_PATH}")
        sys.exit(1)

    # 2. Connect to Database
    try:
        conn = psycopg2.connect(DB_URL)
        cursor = conn.cursor()
    except Exception as e:
        print(f"❌ Failed to connect to database: {e}")
        sys.exit(1)

    # 3. Retrieve JFZ Company
    cursor.execute("SELECT id, name, dental_ceiling, dental_coverage FROM \"InsuranceCompany\" WHERE code = %s", (JFZ_CODE,))
    company = cursor.fetchone()
    if not company:
        print(f"❌ Company with code '{JFZ_CODE}' not found in database.")
        conn.close()
        sys.exit(1)

    company_id, company_name, current_ceiling, current_coverage = company
    print(f"✅ Found Company: {company_name} (ID: {company_id})")
    print(f"    Current Ceiling: {current_ceiling} | Current Coverage: {current_coverage}")

    # Update company ceiling & coverage if not already set or different
    if current_ceiling != DENTAL_CEILING or current_coverage != DENTAL_COV:
        if apply:
            cursor.execute(
                "UPDATE \"InsuranceCompany\" SET dental_ceiling = %s, dental_coverage = %s WHERE id = %s",
                (DENTAL_CEILING, DENTAL_COV, company_id)
            )
            print(f"    ✓ Updated company dental ceiling to {DENTAL_CEILING} and coverage to {DENTAL_COV}%")
        else:
            print(f"    [Dry Run] Will update company dental ceiling to {DENTAL_CEILING} and coverage to {DENTAL_COV}%")

    # 4. Read Excel Rows using openpyxl
    print("\n📂 Reading Excel file...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb.active
        print(f"    Sheet name: {sheet.title}")
    except Exception as e:
        print(f"❌ Failed to read Excel file: {e}")
        conn.close()
        sys.exit(1)

    rows = []
    # Row 1 is header, Row 2 is empty/None in Gelyana file.
    for r in range(3, sheet.max_row + 1):
        name = sheet.cell(row=r, column=1).value
        workplace = sheet.cell(row=r, column=2).value
        primary_card = sheet.cell(row=r, column=3).value
        relation = sheet.cell(row=r, column=4).value
        card_number = sheet.cell(row=r, column=5).value

        # Normalize values
        name = str(name).strip() if name is not None else ""
        workplace = str(workplace).strip() if workplace is not None else ""
        primary_card = str(primary_card).strip().upper() if primary_card is not None else ""
        relation = str(relation).strip().upper() if relation is not None else ""
        card_number = str(card_number).strip().upper() if card_number is not None else ""

        if not name and not card_number:
            continue

        rows.append({
            "row_num": r,
            "name": name,
            "workplace": workplace,
            "primary_card": primary_card,
            "relation": relation,
            "card_number": card_number,
            "is_dependent": not workplace and bool(primary_card)
        })

    print(f"    Total parsed rows from Excel: {len(rows)}")
    primary_count = sum(1 for r in rows if not r["is_dependent"])
    dependent_count = sum(1 for r in rows if r["is_dependent"])
    print(f"    Primary Beneficiaries: {primary_count}")
    print(f"    Dependent Beneficiaries: {dependent_count}")

    # 5. Fetch existing beneficiaries to avoid duplicates
    card_numbers = [r["card_number"] for r in rows if r["card_number"]]
    if card_numbers:
        cursor.execute(
            "SELECT id, card_number, name, deleted_at FROM \"Beneficiary\" WHERE card_number = ANY(%s)",
            (card_numbers,)
        )
        existing_list = cursor.fetchall()
    else:
        existing_list = []

    # Map card_number (uppercase) to its details
    existing_map = {row[1].strip().upper(): {"id": row[0], "name": row[2], "deleted_at": row[3]} for row in existing_list}

    # 6. Process imports
    created_count = 0
    reactivated_count = 0
    skipped_count = 0
    failed_count = 0

    preview = []

    for r in rows:
        card = r["card_number"]
        name = r["name"]
        relation = r["relation"] or "رئيسي"

        if not card:
            print(f"    ⚠️ Row {r['row_num']}: Missing card number for patient '{name}'. Skipping.")
            skipped_count += 1
            continue
        if not name:
            print(f"    ⚠️ Row {r['row_num']}: Missing name for card '{card}'. Skipping.")
            skipped_count += 1
            continue

        if card in existing_map:
            exist = existing_map[card]
            if exist["deleted_at"] is None:
                # Existing and active
                skipped_count += 1
                continue
            else:
                # Deleted, needs restoration
                reactivated_count += 1
                preview.append(("RESTORE", card, name, relation))
                if apply:
                    try:
                        cursor.execute("SAVEPOINT restore_savepoint")
                        cursor.execute(
                            "UPDATE \"Beneficiary\" SET deleted_at = NULL, name = %s, status = 'ACTIVE', total_balance = %s, remaining_balance = %s WHERE id = %s",
                            (name, DENTAL_CEILING, DENTAL_CEILING, exist["id"])
                        )
                        cursor.execute("RELEASE SAVEPOINT restore_savepoint")
                        conn.commit()
                    except Exception as e:
                        print(f"    ❌ Failed to restore {card}: {e}")
                        cursor.execute("ROLLBACK TO SAVEPOINT restore_savepoint")
                        failed_count += 1
                        continue
        else:
            # New beneficiary
            created_count += 1
            preview.append(("CREATE", card, name, relation))
            if apply:
                try:
                    # Use a SAVEPOINT so that if this single row fails, we can rollback only to this savepoint
                    cursor.execute("SAVEPOINT row_savepoint")
                    new_id = generate_cuid()
                    cursor.execute(
                        "INSERT INTO \"Beneficiary\" (id, card_number, name, company_id, status, total_balance, remaining_balance, created_at, failed_attempts, is_legacy_card) VALUES (%s, %s, %s, %s, 'ACTIVE', %s, %s, NOW(), 0, FALSE)",
                        (new_id, card, name, company_id, DENTAL_CEILING, DENTAL_CEILING)
                    )

                    # Create WalletConsumption
                    wallet_id = generate_cuid()
                    cursor.execute(
                        "INSERT INTO \"WalletConsumption\" (id, beneficiary_id, company_id, wallet_type, fiscal_year, consumed_amount, version, created_at, updated_at) VALUES (%s, %s, %s, 'DENTAL', 2026, 0.00, 1, NOW(), NOW()) ON CONFLICT (beneficiary_id, company_id, wallet_type, fiscal_year) DO NOTHING",
                        (wallet_id, new_id, company_id)
                    )
                    cursor.execute("RELEASE SAVEPOINT row_savepoint")
                    # Commit row immediately so we preserve active ones
                    conn.commit()
                except Exception as e:
                    print(f"    ❌ Failed to insert {card}: {e}")
                    cursor.execute("ROLLBACK TO SAVEPOINT row_savepoint")
                    failed_count += 1
                    continue

    # Commit transactions if apply is true
    if apply:
        conn.commit()
        print("\n🏁 Database import committed successfully!")
    else:
        print("\n🏁 Dry run complete. No database changes were made.")

    print("\n" + "=" * 60)
    print("📊 IMPORT STATISTICS:")
    print(f"    Created (New):           {created_count}")
    print(f"    Reactivated (Restored):  {reactivated_count}")
    print(f"    Skipped (Already Active): {skipped_count}")
    if failed_count > 0:
        print(f"    Failed (Errors):         {failed_count}")

    if not apply and preview:
        print("\n📋 Preview of first 20 operations:")
        print(f"{'ACTION':<10} | {'CARD NUMBER':<18} | {'BENEFICIARY NAME':<35} | {'RELATION':<10}")
        print("-" * 80)
        for act, crd, nm, rel in preview[:20]:
            print(f"{act:<10} | {crd:<18} | {nm:<35} | {rel:<10}")
        if len(preview) > 20:
            print(f"    ... and {len(preview) - 20} more actions.")

        print("\n💡 To execute this import in the database, run:")
        print("    python scripts/import_gelyana_beneficiaries.py --apply")
    print("=" * 60)

    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
