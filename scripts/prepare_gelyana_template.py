import os
import sys
import openpyxl

INPUT_PATH = r"اسماء شركات الاسنان\جليانة_علاقات_دقيقة.xlsx"
OUTPUT_PATH = r"اسماء شركات الاسنان\جليانة_جاهز_للاستيراد.xlsx"

def main():
    print("=" * 60)
    print("🛠️  PREPARING GELYANA IMPORT TEMPLATE")
    print(f"    Input:  {INPUT_PATH}")
    print(f"    Output: {OUTPUT_PATH}")
    print("=" * 60)

    # 1. Verify Input File
    if not os.path.exists(INPUT_PATH):
        print(f"❌ Input file not found at: {INPUT_PATH}")
        sys.exit(1)

    # 2. Read Excel
    print("📂 Reading raw Gelyana Excel...")
    try:
        wb_in = openpyxl.load_workbook(INPUT_PATH, data_only=True)
        sheet_in = wb_in.active
        print(f"    Active Sheet: {sheet_in.title}")
    except Exception as e:
        print(f"❌ Failed to load input workbook: {e}")
        sys.exit(1)

    # Common male first names to distinguish Husband (H1) from Wife (W1)
    male_prefixes = {
        "محمد", "أحمد", "احمد", "علي", "على", "حسن", "حسين", "فرج", "محمود", "مصطفى",
        "عمر", "خالد", "صالح", "سعد", "جمعة", "جمعه", "سالم", "عادل", "ابراهيم", "إبراهيم",
        "سليمان", "عبدالسلام", "عبد السلام", "عبدالرحمن", "عبد الرحمن", "عاطف", "هادي",
        "الهادي", "عيسى", "مفتاح", "شعيب", "فضيل", "المبروك", "عاشور", "جبريل", "رجب",
        "فضل", "مسعود", "طه", "ياسين", "خليل", "خميس", "طارق", "شعبان", "ونيس", "ميلاد",
        "حمد", "موسى", "أنور", "يونس", "سراج", "عثمان", "فتحي", "فوزي", "راضي", "رمضان",
        "رياض", "أشرف", "اشرف", "أيمن", "ايمن", "منير", "نبيل", "أسامة", "اسامه", "سفيان",
        "وليد", "حميد", "عبدالحميد", "عبد الحميد", "عبدالعزيز", "عبد العزيز", "عبدالله",
        "عبد الله", "عوض", "عقيلة", "عقيله", "بشير", "ناجي", "امبارك", "الهاشمي", "قطيش",
        "ابوبكر", "أبوبكر", "شريف", "عبدالرحيم", "عبد الرحيم", "أنيس", "انيس", "مختار",
        "علاء", "جمال", "كمال", "سعيد", "سامر", "ساهر", "ماهر", "منذر", "نضال", "عماد",
        "فارس", "معاذ", "زياد", "مروان", "عامر", "حازم", "حاتم", "باسل", "صفوان", "غسان",
        "همام", "هشام", "عصام", "أكرم", "اكرم", "أمجد", "امجد", "سامح", "وائل", "رائد",
        "بهاء", "ضياء", "بليغ", "تامر", "شادي", "هاني", "فادي", "خلف", "جابر", "سليم",
        "منصور", "عباس", "هاشم", "جعفر", "صقر", "فهد", "ذئب", "نمر", "ليث", "أسد",
        "أنس", "انس", "معتز", "حمزة", "حمزه", "طلحة", "طلحه", "قتيبة", "قتيبه", "حذيفة",
        "حذيفه", "مصعب", "بلال", "صابر", "شاكر", "حامد", "محسن", "حليم", "كريم", "لطيف",
        "عبدالسلام", "عبد المطلب", "عبدالملك", "عبد الملك", "عبدالمجيد", "عبد المجيد",
        "عبدالقادر", "عبد القادر", "عبدالحفيظ", "عبد الحفيظ", "عبداللطيف", "عبد اللطيف",
        "عبدالباسط", "عبد الباسط", "عبدالرزاق", "عبد الرزاق", "عبدالحكيم", "عبد الحكيم",
        "عبدالخالق", "عبد الخالق", "عبدالكريم", "عبد الكريم", "عبدالحليم", "عبد الحليم",
        "عبدالوهاب", "عبد الوهاب", "عبدالفتاح", "عبد الفتاح", "عبدالمنعم", "عبد المنعم",
        "عبدالحنان", "عبد الحنان", "عبدالمنان", "عبد المنان", "عبدالمجيد", "عبدالباري",
        "عبد الباري", "عبدالعظيم", "عبد العظيم", "عبدالغني", "عبد الغني", "عبدالمحسن",
        "عبدالباقي", "عبد الباقي", "عبدالرؤوف", "عبد الرؤوف", "عبدالواحد", "عبد الواحد",
        "عبدالجبار", "عبد الجبار", "عبدالمعيد", "عبد المعيد", "عبدالقدوس", "عبد القدوس",
        "عبدالصبور", "عبد الصبور", "عبدالتواب", "عبد التواب", "عبدالمحيي", "عبد المحيي",
        "عبدالمميت", "عبد المميت", "عبدالحي", "عبد الحي", "عبدالقيوم", "عبد القيوم",
        "عبدالواجد", "عبد الواجد", "عبدالماجد", "عبد الماجد", "عبدالواحد", "عبد الواحد",
        "عبدالأحد", "عبد الأحد", "عبدالصمد", "عبد الصمد", "عبدالقادر", "عبد القادر",
        "عبدالمقتدر", "عبد المقتدر", "عبدالمقدم", "عبد المقدم", "عبدالمؤخر", "عبد المؤخر",
        "عبدالأول", "عبد الأول", "العبد", "عبد", "مختار", "موسى", "علاء", "جمال"
    }

    unique_beneficiaries = {}
    total_parsed = 0
    duplicate_count = 0
    corrected_count = 0

    # Start from row 3 (Row 1 is header, Row 2 is empty/None in Gelyana file)
    for r in range(3, sheet_in.max_row + 1):
        name = sheet_in.cell(row=r, column=1).value
        card = sheet_in.cell(row=r, column=5).value

        # Normalize
        name = " ".join(str(name).split()) if name is not None else ""
        card = str(card).strip().upper() if card is not None else ""

        if not name and not card:
            continue

        total_parsed += 1

        if not card:
            print(f"    ⚠️ Row {r}: Missing card number for '{name}'. Skipping.")
            continue
        if not name:
            print(f"    ⚠️ Row {r}: Missing name for card '{card}'. Skipping.")
            continue

        # Correct H1 to W1 if the name is female
        if card.endswith("H1"):
            first_word = name.split()[0] if name else ""
            if first_word not in male_prefixes:
                # This is a female, correct card suffix to W1
                old_card = card
                card = card[:-2] + "W1"
                corrected_count += 1
                if corrected_count <= 10:
                    print(f"    ✏️ Corrected Row {r}: {name} ({old_card} ➔ {card})")

        if card in unique_beneficiaries:
            duplicate_count += 1
            # If the current name is longer or cleaner, we can keep it, but normally we just skip the duplicate card
            continue

        unique_beneficiaries[card] = name

    print(f"    Total processed rows: {total_parsed}")
    print(f"    Corrected H1 to W1 for females: {corrected_count}")
    print(f"    Deduplicated cards found: {duplicate_count}")
    print(f"    Pristine records to export: {len(unique_beneficiaries)}")

    # 3. Create Standardized Output Excel
    print("\n✍️ Writing to standardized Excel template...")
    try:
        wb_out = openpyxl.Workbook()
        sheet_out = wb_out.active
        sheet_out.title = "المستفيدون"

        # Force right-to-left layout for Arabic spreadsheet elegance
        sheet_out.views.sheetView[0].showGridLines = True
        sheet_out.sheet_view.rightToLeft = True

        # Header row
        sheet_out.cell(row=1, column=1, value="الاسم")
        sheet_out.cell(row=1, column=2, value="رقم البطاقة")

        # Apply header formatting
        header_font = openpyxl.styles.Font(name="Arial", size=12, bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        for col in [1, 2]:
            cell = sheet_out.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write data rows
        row_font = openpyxl.styles.Font(name="Arial", size=11)
        row_align_left = openpyxl.styles.Alignment(horizontal="left", vertical="center")
        row_align_center = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        current_row = 2
        for card, name in unique_beneficiaries.items():
            name_cell = sheet_out.cell(row=current_row, column=1, value=name)
            card_cell = sheet_out.cell(row=current_row, column=2, value=card)

            name_cell.font = row_font
            name_cell.alignment = row_align_left

            card_cell.font = row_font
            card_cell.alignment = row_align_center

            current_row += 1

        # Adjust column widths automatically
        sheet_out.column_dimensions['A'].width = 35
        sheet_out.column_dimensions['B'].width = 25

        # Save workbook
        wb_out.save(OUTPUT_PATH)
        print(f"✅ Successfully created standardized template: {OUTPUT_PATH}")
    except Exception as e:
        print(f"❌ Failed to write output workbook: {e}")
        sys.exit(1)

    print("=" * 60)

if __name__ == "__main__":
    main()
